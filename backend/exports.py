"""Excel + PDF export helpers for module data tables."""
import io
import os
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage

from pdf_watermark import attach_watermark

# Brand asset (kept at repo path so both Excel/PDF can embed it)
BRAND_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "itl-logo.jpg")
BRAND_NAME = "INDIAN TRADE LINKS"
BRAND_SUB = "Industrial Services Pvt. Ltd."


# Per-resource column definitions for exports.
COLUMNS: Dict[str, List[Dict[str, str]]] = {
    "clients": [
        {"key": "code", "label": "Code"}, {"key": "name", "label": "Client"},
        {"key": "contact", "label": "Contact"}, {"key": "phone", "label": "Phone"},
        {"key": "gst", "label": "GST"}, {"key": "credit_limit", "label": "Credit Limit"},
        {"key": "status", "label": "Status"},
    ],
    "vendors": [
        {"key": "code", "label": "Code"}, {"key": "name", "label": "Vendor"},
        {"key": "category", "label": "Category"}, {"key": "contact", "label": "Contact"},
        {"key": "phone", "label": "Phone"}, {"key": "rating", "label": "Rating"},
        {"key": "status", "label": "Status"},
    ],
    "employees": [
        {"key": "emp_code", "label": "Code"}, {"key": "name", "label": "Name"},
        {"key": "role", "label": "Role"}, {"key": "department", "label": "Department"},
        {"key": "phone", "label": "Phone"}, {"key": "joining_date", "label": "Joined"},
        {"key": "salary", "label": "Salary"}, {"key": "status", "label": "Status"},
    ],
    "attendance": [
        {"key": "date", "label": "Date"}, {"key": "employee_name", "label": "Employee"},
        {"key": "check_in", "label": "In"}, {"key": "check_out", "label": "Out"},
        {"key": "hours", "label": "Hours"}, {"key": "status", "label": "Status"},
    ],
    "projects": [
        {"key": "code", "label": "Code"}, {"key": "name", "label": "Project"},
        {"key": "client", "label": "Client"}, {"key": "type", "label": "Type"},
        {"key": "site", "label": "Site"}, {"key": "manager", "label": "Manager"},
        {"key": "budget", "label": "Budget"}, {"key": "progress", "label": "Progress %"},
        {"key": "status", "label": "Status"},
    ],
    "inventory": [
        {"key": "code", "label": "Code"}, {"key": "name", "label": "Item"},
        {"key": "category", "label": "Category"}, {"key": "uom", "label": "UOM"},
        {"key": "quantity", "label": "Qty"}, {"key": "min_stock", "label": "Min"},
        {"key": "rate", "label": "Rate"}, {"key": "location", "label": "Location"},
    ],
    "purchase_orders": [
        {"key": "po_number", "label": "PO #"}, {"key": "vendor", "label": "Vendor"},
        {"key": "project", "label": "Project"}, {"key": "date", "label": "Date"},
        {"key": "total", "label": "Total"}, {"key": "paid", "label": "Paid"},
        {"key": "status", "label": "Status"},
    ],
    "quotations": [
        {"key": "quote_number", "label": "Quote #"}, {"key": "client", "label": "Client"},
        {"key": "project", "label": "Project"}, {"key": "date", "label": "Date"},
        {"key": "valid_until", "label": "Valid Till"}, {"key": "total", "label": "Total"},
        {"key": "status", "label": "Status"},
    ],
    "journal_entries": [
        {"key": "je_number", "label": "JE #"}, {"key": "date", "label": "Date"},
        {"key": "account", "label": "Account"}, {"key": "type", "label": "Type"},
        {"key": "cost_centre", "label": "Cost Centre"}, {"key": "amount", "label": "Amount"},
        {"key": "narration", "label": "Narration"},
    ],
    "safety_reports": [
        {"key": "report_id", "label": "Report"}, {"key": "date", "label": "Date"},
        {"key": "project", "label": "Project"}, {"key": "type", "label": "Type"},
        {"key": "severity", "label": "Severity"}, {"key": "reporter", "label": "Reporter"},
        {"key": "status", "label": "Status"},
    ],
    "assets": [
        {"key": "asset_id", "label": "Asset ID"}, {"key": "name", "label": "Asset"},
        {"key": "category", "label": "Category"}, {"key": "purchase_date", "label": "Purchased"},
        {"key": "cost", "label": "Cost"}, {"key": "location", "label": "Location"},
        {"key": "assigned_to", "label": "Assigned"}, {"key": "status", "label": "Status"},
    ],
    "payroll": [
        {"key": "employee_name", "label": "Employee"}, {"key": "month", "label": "Month"},
        {"key": "gross", "label": "Gross"}, {"key": "deductions", "label": "Deductions"},
        {"key": "net", "label": "Net Pay"}, {"key": "status", "label": "Status"},
    ],
    "vehicles": [
        {"key": "reg_number", "label": "Reg #"}, {"key": "type", "label": "Type"},
        {"key": "capacity", "label": "Capacity"}, {"key": "driver", "label": "Driver"},
        {"key": "last_service", "label": "Last Service"}, {"key": "fuel_avg", "label": "Fuel km/L"},
        {"key": "status", "label": "Status"},
    ],
    "documents": [
        {"key": "doc_id", "label": "Doc ID"}, {"key": "title", "label": "Title"},
        {"key": "category", "label": "Category"}, {"key": "project", "label": "Project"},
        {"key": "uploaded_by", "label": "Uploaded By"}, {"key": "version", "label": "Version"},
        {"key": "expiry", "label": "Expiry"},
    ],
    "approvals": [
        {"key": "title", "label": "Request"}, {"key": "type", "label": "Type"},
        {"key": "reference", "label": "Ref"}, {"key": "amount", "label": "Amount"},
        {"key": "requested_by", "label": "Requested By"}, {"key": "status", "label": "Status"},
    ],
}


def to_excel(resource: str, rows: List[Dict[str, Any]]) -> bytes:
    cols = COLUMNS.get(resource, [{"key": k, "label": k} for k in (rows[0].keys() if rows else [])])
    wb = Workbook()
    ws = wb.active
    ws.title = resource[:31]

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Logo (top-left A1:B3)
    try:
        if os.path.exists(BRAND_LOGO_PATH):
            img = XLImage(BRAND_LOGO_PATH)
            img.height = 60
            img.width = 60
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 22
            ws.row_dimensions[3].height = 22
    except Exception:
        pass

    # Brand text block (rows 1-3, columns C+)
    ws.cell(row=1, column=3, value=BRAND_NAME).font = Font(bold=True, size=14, color="0F172A")
    ws.cell(row=2, column=3, value=BRAND_SUB).font = Font(italic=True, color="475569", size=9)
    ws.cell(row=3, column=3, value=f"{resource.replace('_', ' ').title()} · Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}").font = Font(italic=True, color="64748B", size=9)

    # Header row at row 5
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=5, column=j, value=c["label"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for i, row in enumerate(rows, start=6):
        for j, c in enumerate(cols, start=1):
            val = row.get(c["key"])
            if isinstance(val, (dict, list)):
                val = str(val)
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for j, c in enumerate(cols, start=1):
        ws.column_dimensions[ws.cell(row=5, column=j).column_letter].width = max(14, min(40, len(str(c["label"])) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def to_pdf(resource: str, rows: List[Dict[str, Any]]) -> bytes:
    cols = COLUMNS.get(resource, [{"key": k, "label": k} for k in (rows[0].keys() if rows else [])])
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    attach_watermark(doc, width_mm=180.0)  # landscape A4 → wider watermark
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#0f172a"), spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#475569"), spaceAfter=2)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#475569"))

    story = []
    # Top brand band: logo (left) + INDIAN TRADE LINKS title block
    logo_cell = ""
    try:
        if os.path.exists(BRAND_LOGO_PATH):
            logo_cell = RLImage(BRAND_LOGO_PATH, width=20 * mm, height=20 * mm)
    except Exception:
        logo_cell = ""
    brand_block = [
        Paragraph(f"<b>{BRAND_NAME}</b>", title_style),
        Paragraph(BRAND_SUB, sub_style),
        Paragraph(f"{resource.replace('_', ' ').title()} · Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} · {len(rows)} records", meta_style),
    ]
    brand_band = Table([[logo_cell, brand_block]], colWidths=[24 * mm, 240 * mm])
    brand_band.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
        ("RIGHTPADDING", (0, 0), (0, 0), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(brand_band)
    story.append(Spacer(1, 4 * mm))

    table_data = [[c["label"] for c in cols]]
    for r in rows:
        table_data.append([_fmt(r.get(c["key"])) for c in cols])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#ffffff")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (dict, list)):
        return str(v)
    s = str(v)
    return s[:80] + "…" if len(s) > 81 else s
